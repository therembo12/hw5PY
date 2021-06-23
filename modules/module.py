import json
import csv
import os
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime as date


class GetInfo:

    def __init__(self, url):
        self.url = url

    def get_info(self):
        succes_file = open('files/access.log', 'a')
        error_file = open('files/error.log', 'a')
        try:

            response = requests.get(self.url)
            data = response.json()

            succes_file.write(
                f" | {os.getlogin()}|{date.today().strftime('%d.%m.%Y %H:%M:%S')}|\n\
                    {str(response.status_code)} {str(response.reason)} Operation success!\n")

            return (data)

        except Exception as ex:
            error_file.write(
                f" | {os.getlogin()}|{date.today().strftime('%d.%m.%Y %H:%M:%S')}|\n \
                    {str(response.status_code)} {str(response.reason)} Operation failed!\n")
        finally:
            succes_file.close()
            error_file.close()


def write_json(url, content):

    with open(url, 'w') as outfile:
        json.dump(content, outfile, indent=4)
        outfile.close()


def write_csv(url, content):
    header = list(content['exchangeRate'][0].keys())
    with open(url, 'w', encoding="UTF-8", newline='') as file:
        writer = csv.DictWriter(file, fieldnames=header)
        writer.writeheader()
        writer.writerows(content['exchangeRate'])

    file.close()


def write_xlsx(url, content):
    workbook = Workbook()
    sheet = workbook.active
    for index, item in enumerate(list(content['exchangeRate'][0].keys())):
        sheet.cell(row=1, column=index+1).value = item
    row_num = 1
    for row in content['exchangeRate']:
        row_num += 1
        for index, key in enumerate(list(row.keys())):
            sheet.cell(row=row_num, column=index+1).value = row[key]
        workbook.save(filename=url)


def read_json(url):
    try:
        file = open(url, 'r')
        data = json.loads(file.read())
        return data
    finally:
        file.close()


def read_csv(url):
    try:
        with open(url, 'r', encoding="UTF-8", newline='') as file:
            filereader = csv.DictReader(file)
            return filereader
    finally:
        file.close()


def read_xlsx(url):
    try:
        workbook = load_workbook(filename=url)
        workbook.sheetnames

        sheet = workbook.active

        local_data = {}
        local_data['exchangeRate'] = []
        keys = []
        for value in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6, values_only=True):
            for id, key in enumerate(value):
                keys.append(key)

        for values in sheet.iter_rows(min_row=2,
                                      max_row=4,
                                      min_col=1,
                                      max_col=6,
                                      values_only=True):
            row = {}
            for id, value in enumerate(values):
                row.update({keys[id]: value})
            local_data['exchangeRate'].append(row)

        return local_data

    finally:
        workbook.close()
